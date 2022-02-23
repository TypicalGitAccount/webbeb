from openpyxl import load_workbook
from os import listdir, path
from sqlite3 import connect
from datetime import datetime, timedelta, time


class DataRecovery:
    @classmethod
    def recover_gaps(cls, worksheet, col):
        higher_val = None
        lower_val = None
        chlen = len(list(worksheet.rows))
        for row in range(2, chlen+1):
            current = worksheet[f'{col}{row}'].value
            if current is not None:
                higher_val = current
            else:
                missing_len = 0
                for k in range(row, chlen+1):
                    cell = worksheet[f'{col}{k}'].value
                    if cell is not None:
                        lower_val = cell
                        break
                    else:
                        missing_len += 1
                if higher_val is None and lower_val is None:
                    break
                elif higher_val is not None and lower_val is not None:
                    for j in range(row, row + missing_len // 2):
                        worksheet[f'{col}{j}'] = higher_val
                    for j in range(row + missing_len // 2, row + missing_len):
                        worksheet[f'{col}{j}'] = lower_val
                elif higher_val is None:
                    for j in range(row, row + missing_len):
                        worksheet[f'{col}{j}'] = lower_val
                else:
                    for j in range(row, row + missing_len):
                        worksheet[f'{col}{j}'] = higher_val
                row += missing_len

    @classmethod
    def clear_time_column(cls, ws):
        row_ind = 2
        for row in ws.iter_rows(min_row=2):
            row = [cell.value for cell in row]
            if row[1] is None:
                ws.delete_rows(row_ind)
            else:
                row_ind += 1

    @classmethod
    def translate(cls, ws):
        chlen = len(list(ws.rows))
        for row in range(2, chlen + 1):
            if ws[f'D{row}'].value == 'Западный':
                ws[f'D{row}'] = 'Західний'
            if ws[f'D{row}'].value == 'Ю-З':
                ws[f'D{row}'] = 'Південно-західний'
            if ws[f'D{row}'].value == 'С-З':
                ws[f'D{row}'] = 'Північно-західний'
            if ws[f'D{row}'].value == 'Северный':
                ws[f'D{row}'] = 'Північний'
            if ws[f'D{row}'].value == 'Переменный':
                ws[f'D{row}'] = 'Змінний'
            if ws[f'D{row}'].value == 'Южный':
                ws[f'D{row}'] = 'Південний'
            if ws[f'D{row}'].value == 'Ю-В':
                ws[f'D{row}'] = 'Південно-східний'
            if ws[f'D{row}'].value == 'С-В':
                ws[f'D{row}'] = 'Північно-східний'
            if ws[f'D{row}'].value == 'Восточный':
                ws[f'D{row}'].value = 'Східний'

    @classmethod
    def recover_lines(cls, cursor, month, year, city):
        cursor.execute(f'SELECT day, time, temp, wind_direction, wind_speed, month FROM weather WHERE month = {month} AND year = {year} AND city = \'{city}\' ORDER BY day, time')
        lines = cursor.fetchall()

        start_time = datetime.combine(datetime.now().date(), time(hour=0, minute=0))
        end_time = datetime.combine(datetime.now().date(), time(hour=23, minute=30))
        diff = timedelta(minutes=30)
        days = lines[-1][0]

        for day in range(1, int(days) + 1):
            day_lines = [line for line in lines if day == line[0]]
            cur_time = start_time
            line_ind = 0
            while cur_time <= end_time:
                if line_ind >= len(day_lines) or str(cur_time.time()) != day_lines[line_ind][1]:
                    recovered = list(lines[line_ind-1])
                    recovered[0] = day
                    recovered[1] = str(cur_time.time())
                    cursor.execute(
                        f'INSERT INTO weather (day, time, temp, wind_direction, wind_speed, month, year, city) VALUES (?, ?, ?, ?, ?, ?, {year}, \'{city}\')',
                        recovered)
                else:
                    line_ind += 1
                cur_time += diff

    @classmethod
    def recover_files(cls, dir_name, dbname):
        connection = connect(database=dbname)
        cursor = connection.cursor()
        cursor.execute('SELECT * FROM unrecovered')
        unrecovered = cursor.fetchall()

        for file in unrecovered:
            year = file[0].split('-')[1]
            city = file[0].split('-')[0]
            month = file[0].split('-')[-1].split('.')[0]
            filename = path.join(dir_name, file[0])
            wb = load_workbook(filename)
            ws = wb.active
            DataRecovery.translate(ws)
            wb.save(filename)
            DataRecovery.clear_time_column(ws)
            wb.save(filename)
            for col in "ACDE":
                DataRecovery.recover_gaps(ws, col)
            wb.save(filename)
            write_to_db(cursor, file[0], month, year, city)
            connection.commit()
            DataRecovery.recover_lines(cursor, month, year, city)
            connection.commit()
            from_db_to_file(cursor, ws, month, year, city)
            wb.save(filename)
            wb.close()
        cursor.execute("DELETE FROM unrecovered")
        connection.commit()
        connection.close()


def write_to_db(cursor, filename, month, year, city):
    wb = load_workbook(path.dirname(path.abspath(__file__)) + '/exel/' + filename)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        row = list(row)
        row[1] = time_to_str(row[1])
        if row[1] != 'None':
            cursor.execute(
                f'INSERT INTO weather (month, day, time, temp, wind_direction, wind_speed, year, city) VALUES ({month}, ?, ?, ?, ?, ?, {year}, \'{city}\')', row)


def from_db_to_file(cursor, ws, month, year, city):
    ws.delete_rows(2, ws.max_row-1)
    cursor.execute(f'SELECT day, time, temp, wind_direction, wind_speed FROM weather WHERE month = {month} AND year = {year} AND city = \'{city}\' ORDER BY day, time')
    content = cursor.fetchall()
    for line in content:
        ws.append(line)


def time_to_str(val):
    return str(val)
